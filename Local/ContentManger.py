import pandas as pd
import os
import xlsxwriter
from subprocess import Popen
import openpyxl

class ContentManger():
    def __init__(self):
        self.location = os.path.dirname(os.path.realpath(__file__)).replace("\\","/")+"/Files/"

    def GetStoreForcast(self):
        DF_demand = pd.read_csv(self.location+"/"+"store_forecast.csv")
        return DF_demand

    #def EditDemandForcast(self):
    #    path = os.path.realpath(self.location)
    #    os.startfile(path)

    def EditLocalStorClusterParameters(self):
        Popen(self.location + "cluster_parameter.csv", shell=True)
        Popen(self.location + "store_parameter.csv", shell=True)

    def EditStoreForecast(self):
        Popen(self.location+"/"+"store_forecast.csv", shell=True)

    def GetLocalClusterParameters(self):
        location = self.location
        DF_parameters_config = pd.read_csv(location + "cluster_parameter.csv")
        # adjust format of columns
        DF_parameters_config['shift_hour_min'] = DF_parameters_config['shift_hour_min'].astype(int)
        DF_parameters_config['shift_hour_max'] = DF_parameters_config['shift_hour_max'].astype(int)
        DF_parameters_config['driver_hour_min'] = DF_parameters_config['driver_hour_min'].astype(int)
        DF_parameters_config['driver_hour_max'] = DF_parameters_config['driver_hour_max'].astype(int)

        return DF_parameters_config


    def GetLocalStoreParameters(self):
        location = self.location
        DF_parameters_config = pd.read_csv(location + "store_parameter.csv")
        return DF_parameters_config

    def CreateStoreForecastExtract(self, input_DF_store_forecast):
        input_DF_store_forecast.to_csv(self.location+"store_forecasts.csv", index=False)
        open(self.location+"store_forecasts.csv")


    def CreateSchedulesMenue(self, input_DF_clusters):
        input_DF_clusters.to_csv(self.location+"clusters_to_solve.csv", index=False)
        Popen(self.location + "clusters_to_solve.csv", shell=True)


    def CreateExcelReport(self, input_cluster_id, input_DF_schedule=None, input_DF_store_forecaste= None, input_DF_order_forecaste= None, input_DF_driver_forecast = None, input_DF_shifts = None, input_DF_driver_shifts = None, count=None):
        # Create a Pandas Excel writer using XlsxWriter as the engine.

        print("count is " + str(count))
        if count == 0:
            writer = pd.ExcelWriter(self.location+'result/'+'Cluster All result.xlsx', engine='xlsxwriter')

            input_DF_schedule["cluster_id"] = input_cluster_id
            input_DF_store_forecaste["cluster_id"] = input_cluster_id
            input_DF_order_forecaste["cluster_id"] = input_cluster_id
            input_DF_driver_forecast["cluster_id"] = input_cluster_id
            input_DF_shifts["cluster_id"] = input_cluster_id
            input_DF_driver_shifts["cluster_id"] = input_cluster_id

            input_DF_schedule.to_excel(writer, sheet_name='Schedule', index=False)
            input_DF_store_forecaste.to_excel(writer, sheet_name='Store Forecast', index=False)
            input_DF_order_forecaste.to_excel(writer, sheet_name='Order Forecast', index=False)
            input_DF_driver_forecast.to_excel(writer, sheet_name='Driver Forecast', index=False)
            input_DF_shifts.to_excel(writer, sheet_name='Shifts', index=False)
            input_DF_driver_shifts.to_excel(writer, sheet_name='Driver Shift Schedule', index=False)

            writer.save()

        else:

            from openpyxl import load_workbook

            book = load_workbook(self.location+'result/'+'Cluster All result.xlsx')
            writer = pd.ExcelWriter(self.location+'result/'+'Cluster All result.xlsx', engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            # Convert the dataframe to an XlsxWriter Excel object.
            input_DF_schedule["cluster_id"] = input_cluster_id
            input_DF_store_forecaste["cluster_id"] = input_cluster_id
            input_DF_order_forecaste["cluster_id"] = input_cluster_id
            input_DF_driver_forecast["cluster_id"] = input_cluster_id
            input_DF_shifts["cluster_id"] = input_cluster_id
            input_DF_driver_shifts["cluster_id"] = input_cluster_id

            input_DF_schedule.to_excel(writer,sheet_name='Schedule', startrow=writer.sheets['Schedule'].max_row, index = False,header= False)
            input_DF_store_forecaste.to_excel(writer,sheet_name='Store Forecast', startrow=writer.sheets['Store Forecast'].max_row, index = False,header= False)
            input_DF_order_forecaste.to_excel(writer,sheet_name='Order Forecast', startrow=writer.sheets['Order Forecast'].max_row, index = False,header= False)
            input_DF_driver_forecast.to_excel(writer,sheet_name='Driver Forecast', startrow=writer.sheets['Driver Forecast'].max_row, index = False,header= False)
            input_DF_shifts.to_excel(writer,sheet_name='Shifts', startrow=writer.sheets['Shifts'].max_row, index = False,header= False)
            input_DF_driver_shifts.to_excel(writer,sheet_name='Driver Shift Schedule', startrow=writer.sheets['Driver Shift Schedule'].max_row, index = False,header= False)

            #worksheet = writer.sheets['Order Forecast']
            writer.save()
        #Popen(self.location + "Model Result.xlsx", shell=True)

